import os
import sys
import re
import json
import random
import datetime
import openpyxl
import http.server

def clean_str(val):
    if val is None:
        return ""
    return str(val).strip()

def sanitize_filename(name):
    # remove invalid Windows filename characters
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def extract_date_from_filename(filename):
    # Match patterns like "Jul 06, 2026" or "Mar 02, 2026" or "Sep 20, 2024"
    match = re.search(r'([A-Za-z]{3}\s+\d{1,2},\s+\d{4})', filename)
    if match:
        return match.group(1)
    
    # Try dashed/dotted dates like 2026-07-06 or 06.07.2026
    match_iso = re.search(r'(\d{4}-\d{2}-\d{2})', filename)
    if match_iso:
        try:
            dt = datetime.datetime.strptime(match_iso.group(1), "%Y-%m-%d")
            return dt.strftime("%b %d, %Y")
        except ValueError:
            pass
            
    return None

def parse_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    plan_sheets = [name for name in wb.sheetnames if name.lower().startswith("plan")]
    if not plan_sheets:
        if len(wb.sheetnames) > 0:
            sheetname = wb.sheetnames[0]
        else:
            return []
    else:
        sheetname = plan_sheets[0]
    
    sheet = wb[sheetname]
    
    practices = []
    current_practice = None
    current_cycle = None
    
    for r in range(1, sheet.max_row + 1):
        col1 = sheet.cell(r, 1).value
        col2 = sheet.cell(r, 2).value
        col3 = sheet.cell(r, 3).value
        col4 = sheet.cell(r, 4).value
        col7 = sheet.cell(r, 7).value
        
        c1_str = clean_str(col1)
        is_header = c1_str.startswith("Practice Plan") or (c1_str.startswith("Practice") and ":" in c1_str)
        
        if is_header:
            should_merge = (
                current_practice is not None and
                r - current_practice["start_row"] <= 5 and
                not any(c["intervals"] for c in current_practice["cycles"])
            )
            
            if not should_merge:
                if current_practice:
                    practices.append(current_practice)
                current_practice = {
                    "title": "",
                    "plan_header": "",
                    "cycles": [],
                    "start_row": r
                }
                current_cycle = {
                    "name": "Initial",
                    "sets": 1,
                    "intervals": []
                }
                current_practice["cycles"].append(current_cycle)
                
            if c1_str.startswith("Practice Plan"):
                current_practice["plan_header"] = c1_str
            else:
                current_practice["title"] = c1_str
            continue
            
        if current_practice is None:
            continue
            
        if c1_str.startswith("Practice:") or (c1_str.startswith("Practice") and ":" in c1_str):
            current_practice["title"] = c1_str
            continue
            
        if c1_str.startswith('"type":') or c1_str.startswith("type:"):
            m = re.search(r'type["\s:]+(\d)', c1_str)
            if not m:
                continue
            type_idx = int(m.group(1))
            
            duration = col2
            if duration is None:
                continue
            try:
                duration = float(duration)
            except ValueError:
                continue
                
            if duration <= 0:
                continue
                
            unit = clean_str(col3).lower()
            if "min" in unit:
                seconds = int(duration * 60)
            elif "sec" in unit:
                seconds = int(duration)
            else:
                seconds = int(duration * 60)
                
            desc = clean_str(col4)
            notes = clean_str(col7)
            
            interval = {
                "type": type_idx,
                "time": seconds,
                "description": desc,
                "notes": notes
            }
            current_cycle["intervals"].append(interval)
            continue
            
        if col2 is not None:
            try:
                sets_count = float(col2)
                if c1_str != "" and not c1_str.startswith("Type") and not c1_str.startswith("Practice"):
                    current_cycle = {
                        "name": c1_str,
                        "sets": int(sets_count),
                        "intervals": []
                    }
                    current_practice["cycles"].append(current_cycle)
            except ValueError:
                pass
                
        if col1 is None and col2 is None:
            if current_cycle and current_cycle["intervals"]:
                current_cycle = {
                    "name": "Standalone",
                    "sets": 1,
                    "intervals": []
                }
                current_practice["cycles"].append(current_cycle)
                
    if current_practice:
        practices.append(current_practice)
        
    # Clean up empty cycles
    for p in practices:
        p["cycles"] = [c for c in p["cycles"] if c["intervals"]]
        
    return [p for p in practices if p["cycles"]]

def clean_label(label):
    # Standardize string formatting rules:
    # 1. No double quotes
    label = label.replace('"', '')
    # 2. Minutes: Use ' (e.g. 10' Warm-up)
    label = re.sub(r'(\d+)\s*(min|mins|minute|minutes)', r"\1'", label)
    # 3. Seconds: Use sec or seconds
    label = re.sub(r'(\d+)\s*(sec|secs|seconds)', r"\1 sec", label)
    # 4. Rest Indicators: Use R or Rest
    label = re.sub(r'\b(r|rest)\b', 'Rest', label, flags=re.IGNORECASE)
    # 5. Gears: G[Number]
    label = re.sub(r'\bgear\s*(\d+)\b', r"G\1", label, flags=re.IGNORECASE)
    label = re.sub(r'\bg\s*(\d+)\b', r"G\1", label, flags=re.IGNORECASE)
    return label.strip()

def format_interval(inv):
    desc = inv.get("description", "").replace('"', '')
    parts = [
        f'"description": "{desc}"',
        f'"time": {inv.get("time")}',
        f'"type": {inv.get("type")}',
        f'"addSet": {str(inv.get("addSet", False)).lower()}',
        f'"bpm": {inv.get("bpm", 0)}',
        f'"cycle": {inv.get("cycle", -1)}',
        f'"cyclesCount": {inv.get("cyclesCount", -1)}',
        f'"isRepsMode": {str(inv.get("isRepsMode", False)).lower()}',
        f'"reps": {inv.get("reps", 0)}',
        f'"tabata": {inv.get("tabata", -1)}',
        f'"tabatasCount": {inv.get("tabatasCount", -1)}'
    ]
    return "      { " + ", ".join(parts) + " }"

def generate_tabata_json(title, expanded_intervals, prep, work, rest, cooldown, tabatas_count):
    color_id = random.randint(0, 15)
    workout_id = random.randint(100, 999)
    
    # Generate intervals formatting string with single line constraint
    interval_lines = [format_interval(inv) for inv in expanded_intervals]
    intervals_str = ",\n".join(interval_lines)
    
    json_template = f"""{{
  "workout": {{
    "colorId": {color_id},
    "coolDown": {cooldown},
    "cycles": 1,
    "doNotRepeatFirstPrepareAndLastCoolDown": false,
    "id": {workout_id},
    "intervals": [
{intervals_str}
    ],
    "intervalsSetsCount": 1,
    "isFavorite": false,
    "isRestRepsMode": false,
    "isWorkRepsMode": false,
    "prepare": {prep},
    "rest": {rest},
    "restBetweenTabatas": 0,
    "restBetweenWorkoutsReps": 0,
    "restBetweenWorkoutsRepsMode": false,
    "restBetweenWorkoutsTime": 0,
    "restBpm": 0,
    "restDescription": "Rest",
    "restReps": 0,
    "skipLastRestInterval": true,
    "skipPrepareAndCoolDownBetweenWorkouts": false,
    "tabatasCount": {tabatas_count},
    "title": "{title}",
    "work": {work},
    "workBpm": 0,
    "workDescription": "Work",
    "workReps": 0
  }},
  "fileVersion": 1,
  "packageName": "com.evgeniysharafan.tabatatimer",
  "platform": 1,
  "type": 1,
  "versionCode": 502005,
  "versionName": "5.2.5"
}}"""
    return json_template

def format_duration_shorthand(seconds):
    mins = seconds // 60
    secs = seconds % 60
    if secs == 0:
        return f"{mins}'"
    return f"{mins}' {secs}\""

def infer_heuristics(cycles):
    # 1. Generate Sport Shorthand Summary
    summary_parts = []
    all_descriptions_and_notes = []
    
    for c in cycles:
        cycle_name = c["name"]
        sets = c["sets"]
        intervals = c["intervals"]
        
        # build list of descriptions/times for this cycle
        sub_parts = []
        for inv in intervals:
            time_str = format_duration_shorthand(inv["time"])
            desc = inv["description"]
            all_descriptions_and_notes.append(desc.lower())
            if inv["notes"]:
                all_descriptions_and_notes.append(inv["notes"].lower())
                
            # shorthand format
            if "rest" in desc.lower() or inv["type"] == 2:
                sub_parts.append(f"{time_str} Rest")
            else:
                gear_match = re.search(r'\b(G\d+(\.\d+)?)\b', desc, flags=re.IGNORECASE)
                if gear_match:
                    sub_parts.append(f"{time_str} @ {gear_match.group(1).upper()}")
                else:
                    sub_parts.append(f"{time_str} {desc}" if desc else time_str)
        
        cycle_summary = " + ".join(sub_parts)
        if sets > 1:
            summary_parts.append(f"{sets} sets of [{cycle_summary}]")
        else:
            summary_parts.append(cycle_summary)
            
    full_summary = " + ".join(summary_parts)
    
    # 2. Infer Tags
    tags = []
    text_corpus = " ".join(all_descriptions_and_notes)
    
    # Heuristics based on text corpus
    if any(g in text_corpus for g in ["g10", "g9", "g8", "sprint", "hit"]):
        tags.append("HIT")
    if any(g in text_corpus for g in ["g5", "g6", "g7", "2k", "1k", "500m", "race"]):
        tags.append("Race Ready")
    if any(g in text_corpus for g in ["endurance", "aet", "long", "aerobic", "g4", "g3"]):
        tags.append("Endurance")
    if any(keyword in text_corpus for keyword in ["drill", "stroke", "catch", "pull", "exit", "technical", "technique"]):
        tags.append("Technical paddling")
    if any(keyword in text_corpus for keyword in ["recovery", "cooldown", "g0", "easy"]):
        tags.append("Recovery")
        
    if not tags:
        tags.append("Endurance") # default
        
    return full_summary, tags

def clean_title(title_cell, plan_header_cell):
    # Remove prefix "Practice Plan X: " or "Practice: "
    clean_p = re.sub(r'^practice plan[^:]*:\s*', '', plan_header_cell, flags=re.IGNORECASE)
    clean_t = re.sub(r'^practice[^:]*:\s*', '', title_cell, flags=re.IGNORECASE)
    return clean_p.strip(), clean_t.strip()

def update_sessions_json(date, title, summary, tags, sheet_rel_path, tabata_rel_path):
    log_path = "sessions.json"
    data = []
    if os.path.exists(log_path):
        try:
            with open(log_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            data = []
            
    # Check if duplicate entry exists (same date and title)
    existing_idx = -1
    for idx, item in enumerate(data):
        if item.get("date") == date and item.get("title") == title:
            existing_idx = idx
            break
            
    entry = {
        "date": date,
        "title": title,
        "summary": summary,
        "tags": tags,
        "spreadsheet_path": sheet_rel_path,
        "tabata_path": tabata_rel_path
    }
    
    if existing_idx != -1:
        data[existing_idx] = entry
        print(f"[Log] Updated existing entry for '{title}' on {date}.")
    else:
        data.append(entry)
        print(f"[Log] Appended new entry for '{title}' on {date}.")
        
    # Sort chronologically by converting dates
    def get_sort_key(item):
        date_str = item.get("date", "")
        try:
            return datetime.datetime.strptime(date_str, "%b %d, %Y")
        except Exception:
            return datetime.datetime.min
            
    data.sort(key=get_sort_key, reverse=True)
    
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def log_pp_only_mode():
    print("\n=== Log Practice Plan Only Mode ===")
    
    # Prompt for details
    date_input = input("Enter Date (e.g. Jul 06, 2026) [default: Today]: ").strip()
    if not date_input:
        date_input = datetime.datetime.now().strftime("%b %d, %Y")
        
    title_input = input("Enter Workout Title (e.g. Alternate Long Paddle): ").strip()
    while not title_input:
        title_input = input("Title is required! Enter Title: ").strip()
        
    summary_input = input("Enter Workout Summary Text (e.g. 5x [9' @ G4 + 4' Rest]): ").strip()
    
    print("\nAvailable Tags: HIT, Race Ready, Endurance, Technical paddling, Recovery")
    tags_input = input("Enter tags (comma separated, e.g. HIT, Race Ready): ").strip()
    tags = [t.strip() for t in tags_input.split(",") if t.strip()]
    if not tags:
        tags = ["Endurance"]
        
    update_sessions_json(date_input, title_input, summary_input, tags, None, None)
    print("\nSuccessfully logged historical practice plan!")

def run_pipeline(manual_file=None, non_interactive=False):
    print("=== Antigravity Dragon Boat Ingestion Pipeline ===")
    
    # Find files to process
    files_to_process = []
    if manual_file:
        if os.path.exists(manual_file):
            files_to_process.append(manual_file)
        else:
            print(f"[Error] File not found: {manual_file}")
            sys.exit(1)
    else:
        # scan root directory for loose spreadsheets
        for f in os.listdir("."):
            if f.endswith(".xlsx") and not f.startswith("~$"):
                files_to_process.append(f)
                
    if not files_to_process:
        print("[Info] No loose spreadsheets found to process in the root folder.")
        return
        
    os.makedirs("01-Calculators_Sheets", exist_ok=True)
    os.makedirs("02-App_Timers_JSON", exist_ok=True)
    
    for filepath in files_to_process:
        print(f"\nProcessing workbook: {filepath}")
        
        # Parse spreadsheet
        practices = parse_xlsx(filepath)
        if not practices:
            print(f"[Warning] No practices/timing rows parsed from {filepath}. Skipping.")
            continue
            
        # Extract date
        date_extracted = extract_date_from_filename(filepath)
        if not date_extracted:
            # Fallback to file creation time or today
            mtime = os.path.getmtime(filepath)
            date_extracted = datetime.datetime.fromtimestamp(mtime).strftime("%b %d, %Y")
            
        print(f"Extracted Date: {date_extracted}")
        
        for p in practices:
            # Title resolution
            clean_plan, clean_tit = clean_title(p["title"], p["plan_header"])
            
            # Check for conflict
            is_conflict = False
            final_title = clean_tit
            
            if clean_plan and clean_tit:
                if clean_plan.lower() != clean_tit.lower():
                    is_conflict = True
                    # if tit cell is FullB-9minTran (common template leftover), prefer the plan header
                    if "fullb-9mintran" in clean_tit.lower() or "halfb-9mintran" in clean_tit.lower():
                        final_title = clean_plan
                    else:
                        final_title = clean_tit
            elif clean_plan:
                final_title = clean_plan
            elif clean_tit:
                final_title = clean_tit
            else:
                final_title = "Untitled Practice"
                
            # If template default or conflict is found, warn
            if is_conflict or "fullb" in clean_tit.lower() or "halfb" in clean_tit.lower():
                print(f"\n[WARNING] Potential Title Conflict/Leftover detected in sheet:")
                print(f"  - Practice Plan header cell: '{clean_plan}'")
                print(f"  - Practice cell: '{clean_tit}'")
                print(f"  - Inferred Default Title: '{final_title}'")
                
                if not non_interactive:
                    print("Choose title option:")
                    print(f"  [1] {final_title} (Recommended)")
                    if clean_plan and final_title != clean_plan:
                        print(f"  [2] {clean_plan}")
                    if clean_tit and final_title != clean_tit:
                        print(f"  [3] {clean_tit}")
                    print("  [4] Enter custom title")
                    
                    choice = input("Enter choice [1-4]: ").strip()
                    if choice == "2" and clean_plan:
                        final_title = clean_plan
                    elif choice == "3" and clean_tit:
                        final_title = clean_tit
                    elif choice == "4":
                        final_title = input("Enter custom title: ").strip()
                        
            print(f"Final resolved title: '{final_title}'")
            
            # Generate summary and tags
            inferred_summary, inferred_tags = infer_heuristics(p["cycles"])
            
            final_summary = inferred_summary
            final_tags = inferred_tags
            
            if not non_interactive:
                print(f"\nInferred Summary: {inferred_summary}")
                print(f"Inferred Tags: {', '.join(inferred_tags)}")
                
                confirm = input("Would you like to override summary or tags? (y/n) [n]: ").strip().lower()
                if confirm == "y":
                    override_sum = input(f"Enter summary [default: {inferred_summary}]: ").strip()
                    if override_sum:
                        final_summary = override_sum
                    
                    print("Available tags: HIT, Race Ready, Endurance, Technical paddling, Recovery")
                    override_tags = input(f"Enter tags (comma separated) [default: {', '.join(inferred_tags)}]: ").strip()
                    if override_tags:
                        final_tags = [t.strip() for t in override_tags.split(",") if t.strip()]
            
            # Check for matching .tabata file in the same folder as the excel file
            excel_dir = os.path.dirname(filepath) or "."
            excel_filename = os.path.basename(filepath)
            base_name, _ = os.path.splitext(excel_filename)
            matching_tabata_filename = base_name + ".tabata"
            matching_tabata_path = os.path.join(excel_dir, matching_tabata_filename)
            
            associated_tabata_src = None
            if os.path.exists(matching_tabata_path):
                associated_tabata_src = matching_tabata_path
                print(f"[Timer] Found matching Tabata file: {matching_tabata_path}")
            else:
                if not non_interactive:
                    assoc = input(f"No matching .tabata file found for '{excel_filename}'. Do you wish to associate an existing JSON/tabata file? (y/n) [n]: ").strip().lower()
                    if assoc == "y":
                        tabata_file = input("Enter the JSON/tabata file name or path: ").strip()
                        resolved_path = os.path.join(excel_dir, tabata_file)
                        if os.path.exists(tabata_file):
                            associated_tabata_src = tabata_file
                        elif os.path.exists(resolved_path):
                            associated_tabata_src = resolved_path
                        else:
                            print(f"[Warning] Associated file not found: '{tabata_file}'. Proceeding without Tabata association.")
            
            # Destination path naming
            sanitized_title = sanitize_filename(final_title)
            new_sheet_name = f"{date_extracted} - {sanitized_title}.xlsx"
            sheet_dest_path = os.path.join("01-Calculators_Sheets", new_sheet_name)
            
            tabata_dest_rel = None
            if associated_tabata_src:
                new_tabata_name = f"{date_extracted} - {sanitized_title}.tabata"
                tabata_dest_path = os.path.join("02-App_Timers_JSON", new_tabata_name)
                import shutil
                shutil.copy2(associated_tabata_src, tabata_dest_path)
                print(f"[Timer] Copied associated Tabata file to: {tabata_dest_path}")
                tabata_dest_rel = tabata_dest_path.replace("\\", "/")
                
                # Cleanup loose matching tabata file if it was in the root
                if not manual_file and os.path.exists(associated_tabata_src):
                    norm_t = os.path.normpath(associated_tabata_src)
                    if os.sep not in norm_t:
                        try:
                            os.remove(associated_tabata_src)
                            print(f"[Cleanup] Removed loose root Tabata file: {associated_tabata_src}")
                        except Exception as e:
                            pass
            
            # Copy/Move Excel
            import shutil
            shutil.copy2(filepath, sheet_dest_path)
            print(f"[Archive] Copied workbook calculator to: {sheet_dest_path}")
            
            # Update log
            update_sessions_json(
                date_extracted,
                final_title,
                final_summary,
                final_tags,
                sheet_dest_path.replace("\\", "/"),
                tabata_dest_rel
            )
            
        # If it was a loose file in root, remove the original to clean root folder
        if not manual_file and os.path.exists(filepath):
            norm_f = os.path.normpath(filepath)
            if os.sep not in norm_f:
                try:
                    os.remove(filepath)
                    print(f"[Cleanup] Removed loose root spreadsheet: {filepath}")
                except Exception as e:
                    print(f"[Warning] Could not remove loose file {filepath}: {e}")

class DashboardAPIHandler(http.server.SimpleHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        if self.path == '/api/sessions/edit':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                params = json.loads(post_data.decode('utf-8'))
                original_date = params.get('original_date')
                original_title = params.get('original_title')
                new_date = params.get('date')
                new_title = params.get('title')
                new_summary = params.get('summary')
                new_tags = params.get('tags', [])
                
                log_path = "sessions.json"
                data = []
                if os.path.exists(log_path):
                    with open(log_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                
                updated = False
                for item in data:
                    if item.get("date") == original_date and item.get("title") == original_title:
                        item["date"] = new_date
                        item["title"] = new_title
                        item["summary"] = new_summary
                        item["tags"] = new_tags
                        updated = True
                        break
                
                if updated:
                    def get_sort_key(item):
                        date_str = item.get("date", "")
                        try:
                            return datetime.datetime.strptime(date_str, "%b %d, %Y")
                        except Exception:
                            return datetime.datetime.min
                    data.sort(key=get_sort_key, reverse=True)
                    
                    with open(log_path, "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=2, ensure_ascii=False)
                    
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "success"}).encode('utf-8'))
                else:
                    self.send_response(404)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.end_headers()
                    self.wfile.write(json.dumps({"status": "error", "message": "Session not found"}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode('utf-8'))
                
        elif self.path == '/api/run-pipeline':
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            try:
                params = json.loads(post_data.decode('utf-8'))
                cmd_type = params.get('type')
                
                if cmd_type == 'standard':
                    cmd = 'python pipeline.py'
                elif cmd_type == 'log-only':
                    cmd = 'python pipeline.py --log-only'
                elif cmd_type == 'yes':
                    cmd = 'python pipeline.py --yes'
                else:
                    raise ValueError(f"Unknown command type: {cmd_type}")
                
                import subprocess
                subprocess.Popen(
                    ["powershell.exe", "-NoExit", "-Command", cmd],
                    creationflags=subprocess.CREATE_NEW_CONSOLE,
                    cwd=os.getcwd()
                )
                
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "success", "command": cmd}).encode('utf-8'))
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(json.dumps({"status": "error", "message": str(e)}).encode('utf-8'))
        else:
            self.send_response(404)
            self.end_headers()

def start_server():
    import http.server
    import webbrowser
    
    port = 8000
    server = None
    while port < 8010:
        try:
            handler = DashboardAPIHandler
            server = http.server.HTTPServer(("127.0.0.1", port), handler)
            break
        except OSError:
            print(f"[Server] Port {port} is busy. Trying next port...")
            port += 1
            
    if not server:
        print("[Error] Could not find a free port between 8000 and 8010.")
        sys.exit(1)
        
    print(f"\n=======================================================")
    print(f"Dragon Boat Dashboard Server Running at:")
    print(f"http://localhost:{port}")
    print(f"=======================================================\n")
    
    try:
        webbrowser.open(f"http://localhost:{port}")
    except Exception as e:
        print(f"[Warning] Could not open browser automatically: {e}")
        
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[Server] Shutting down server...")
        server.server_close()
        sys.exit(0)

if __name__ == "__main__":
    import http.server
    if len(sys.argv) > 1:
        if "--server" in sys.argv:
            start_server()
        elif "--log-only" in sys.argv or "/log-PP-only" in sys.argv:
            log_pp_only_mode()
        else:
            non_int = "--yes" in sys.argv or "-y" in sys.argv
            manual_file = None
            for arg in sys.argv[1:]:
                if not arg.startswith("-"):
                    manual_file = arg
                    break
            run_pipeline(manual_file=manual_file, non_interactive=non_int)
    else:
        run_pipeline()
